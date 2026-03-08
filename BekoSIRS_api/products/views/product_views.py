# products/views/product_views.py
"""
Product and Category management views.
"""

from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, AllowAny, IsAdminUser
from django.db.models import Count
from django.http import HttpResponse

from products.models import Product, Category, ProductOwnership, WishlistItem, Notification, ProductAssignment
from products.serializers import ProductSerializer, CategorySerializer


class ProductViewSet(viewsets.ModelViewSet):
    """Product CRUD operations with role-based access."""
    queryset = Product.objects.all().select_related("category")
    serializer_class = ProductSerializer

    def get_permissions(self):
        if self.action in ["list", "retrieve", "popular"]:
            return [AllowAny()]
        from products.permissions import IsSeller
        return [IsSeller()]

        return queryset

    def get_queryset(self):
        """
        Filter products by search query and category.
        Search looks in: name, brand, description, category name, model code
        """
        from django.db.models import Q
        from django.utils import timezone
        
        # Import dynamically to avoid circular usage if model is in same app
        from products.models import SearchHistory
        
        queryset = Product.objects.all().select_related("category")
        
        # Search filter
        search_query = self.request.query_params.get('search', None)
        if search_query:
            search_query = search_query.strip()
            queryset = queryset.filter(
                Q(name__icontains=search_query) |
                Q(brand__icontains=search_query) |
                Q(description__icontains=search_query) |
                Q(category__name__icontains=search_query) |
                Q(model_code__icontains=search_query)
            )
            
            # Record search history for logged-in customers
            user = self.request.user
            if user.is_authenticated and hasattr(user, 'role') and user.role == 'customer':
                try:
                    # Avoid duplicate logging if recent search exists (e.g. 5 minutes)
                    last_search = SearchHistory.objects.filter(
                        customer=user, 
                        query__iexact=search_query
                    ).order_by('-created_at').first()
                    
                    if not last_search or (timezone.now() - last_search.created_at).seconds > 300:
                        SearchHistory.objects.create(customer=user, query=search_query)
                except Exception as e:
                    print(f"Search history error: {e}")
        
        # Category filter
        category = self.request.query_params.get('category', None)
        if category:
            queryset = queryset.filter(category_id=category)
        
        return queryset

    @action(
        detail=False,
        methods=["get"],
        url_path="my-products",
        permission_classes=[IsAuthenticated],
    )
    def my_products(self, request):
        """GET /api/products/my-products/ - User's assigned products."""
        user = request.user

        if user.role in ["admin", "seller"]:
            qs = Product.objects.all().select_related("category")
            return Response(ProductSerializer(qs, many=True).data)

        ownerships = (
            ProductOwnership.objects.filter(customer=user)
            .select_related("product", "product__category")
            .order_by("-id")
        )

        result = []
        for o in ownerships:
            p = o.product
            item = ProductSerializer(p).data
            if hasattr(o, "assigned_date"):
                item["assigned_date"] = o.assigned_date
            elif hasattr(o, "assigned_at"):
                item["assigned_date"] = o.assigned_at
            elif hasattr(o, "created_at"):
                item["assigned_date"] = o.created_at
            if hasattr(o, "status"):
                item["status"] = o.status
            result.append(item)

        return Response(result)

    @action(
        detail=False,
        methods=["get"],
        url_path="popular",
        permission_classes=[AllowAny],
    )
    def popular(self, request):
        """GET /api/v1/products/popular/ - Most assigned products."""
        # Count how many times each product has been assigned
        product_counts = (
            ProductAssignment.objects
            .values('product')
            .annotate(assignment_count=Count('id'))
            .order_by('-assignment_count')
        )
        
        # Get product IDs and create a mapping of product_id -> count
        product_id_to_count = {item['product']: item['assignment_count'] for item in product_counts}
        product_ids = list(product_id_to_count.keys())
        
        # Fetch products in the sorted order
        products = Product.objects.filter(id__in=product_ids).select_related('category')
        
        # Sort products by assignment count
        sorted_products = sorted(
            products,
            key=lambda p: product_id_to_count.get(p.id, 0),
            reverse=True
        )
        
        # Serialize and return
        serializer = self.get_serializer(sorted_products, many=True)
        return Response(serializer.data)

    def perform_update(self, serializer):
        """Detect price changes and send notifications."""
        instance = self.get_object()
        old_price = instance.price
        new_price = serializer.validated_data.get('price', old_price)
        updated_instance = serializer.save()

        if new_price and new_price < old_price:
            self._send_price_drop_notifications(updated_instance, old_price, new_price)

    def _send_price_drop_notifications(self, product, old_price, new_price):
        """Send price drop notifications to wishlist users."""
        discount_percent = round((float(old_price) - float(new_price)) / float(old_price) * 100, 1)
        
        wishlist_items = WishlistItem.objects.filter(
            product=product,
            notify_on_price_drop=True
        ).select_related('wishlist__customer')

        notifications = []
        for item in wishlist_items:
            user = item.wishlist.customer
            if user.notify_price_drops:
                notifications.append(
                    Notification(
                        user=user,
                        notification_type='price_drop',
                        title=f'Fiyat Düştü! %{discount_percent} İndirim',
                        message=f'{product.name} ürününün fiyatı {old_price}₺ yerine {new_price}₺ oldu!',
                        related_product=product
                    )
                )

        if notifications:
            Notification.objects.bulk_create(notifications)


class CategoryViewSet(viewsets.ModelViewSet):
    """Category CRUD with product count annotation."""
    queryset = Category.objects.annotate(product_count=Count('products')).all()
    serializer_class = CategorySerializer

    def get_permissions(self):
        if self.action in ["list", "retrieve"]:
            return [AllowAny()]
        return [IsAdminUser()]


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def my_products_direct(request):
    """GET /api/my-products/ - Direct endpoint for mobile compatibility."""
    user = request.user

    if user.role in ["admin", "seller"]:
        qs = Product.objects.all().select_related("category")
        return Response(ProductSerializer(qs, many=True).data)

    ownerships = (
        ProductOwnership.objects.filter(customer=user)
        .select_related("product", "product__category")
        .order_by("-id")
    )

    result = []
    for o in ownerships:
        p = o.product
        item = ProductSerializer(p).data
        if hasattr(o, "assigned_date"):
            item["assigned_date"] = o.assigned_date
        elif hasattr(o, "assigned_at"):
            item["assigned_date"] = o.assigned_at
        elif hasattr(o, "created_at"):
            item["assigned_date"] = o.created_at
        if hasattr(o, "status"):
            item["status"] = o.status
        result.append(item)

    return Response(result)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_products_excel(request):
    """GET /api/products/export/excel/ - Export products as Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    user = request.user
    if user.role not in ['admin', 'seller']:
        return Response({'error': 'Yetkisiz erişim'}, status=status.HTTP_403_FORBIDDEN)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ürünler"

    # Header styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = [
        "ID", "Ürün Adı", "Marka", "Model Kodu", "Kategori",
        "List Fiyatı (₺)", "Peşin Fiyatı (₺)", "Stok", 
        "Garanti (Ay)", "Garanti Kodu", "Kampanya"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    products = Product.objects.all().select_related('category')

    for row, product in enumerate(products, 2):
        ws.cell(row=row, column=1, value=product.id).border = thin_border
        ws.cell(row=row, column=2, value=product.name).border = thin_border
        ws.cell(row=row, column=3, value=product.brand).border = thin_border
        ws.cell(row=row, column=4, value=getattr(product, 'model_code', '')).border = thin_border
        ws.cell(row=row, column=5, value=product.category.name if product.category else '').border = thin_border
        ws.cell(row=row, column=6, value=float(product.price) if product.price else 0).border = thin_border
        ws.cell(row=row, column=7, value=float(getattr(product, 'price_cash', 0) or 0)).border = thin_border
        ws.cell(row=row, column=8, value=product.stock).border = thin_border
        ws.cell(row=row, column=9, value=product.warranty_duration_months).border = thin_border
        ws.cell(row=row, column=10, value=getattr(product, 'warranty_code', '')).border = thin_border
        ws.cell(row=row, column=11, value=getattr(product, 'campaign_tag', '')).border = thin_border

    column_widths = [8, 40, 15, 15, 20, 15, 15, 10, 12, 15, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="bekosirs_products.xlsx"'
    wb.save(response)
    return response
